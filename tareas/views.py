from datetime import timezone
from decimal import Decimal
import json
from django.template.loader import get_template
from pyexpat.errors import messages
from django.shortcuts import get_object_or_404, render, redirect
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from .models import Carrito, Factura, DetalleFactura, Recoleccion, RegistroResiduo, Rol
from django.db.models import Sum, F, FloatField, ExpressionWrapper
from io import BytesIO
from django.template.loader import render_to_string
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.contrib.auth import login
from django.contrib.auth.views import LoginView
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.shortcuts import redirect
from django.contrib.auth import logout
from django.utils.timezone import now
from django.db.models import Sum, F, FloatField, ExpressionWrapper
import stripe
from django.core.serializers.json import DjangoJSONEncoder
from django.http import HttpResponseRedirect, JsonResponse
from .models import CategoriaResiduo, DetalleFactura, Empresa, Factura, Ubicacion, UbicacionCategoria,  RegistroResiduo
from django.shortcuts import render
from .models import Ubicacion, Usuario
from django.conf import settings
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, get_object_or_404
from .models import Evento
from .forms import CategoriaResiduoForm, CotizacionDomicilioForm, EmpresaForm, EventoForm, InformeManejoResiduosForm, InformeNormativasForm, RecoleccionForm, RegistroGestorForm,  RegistroUsuarioForm, UbicacionCategoriaFormSet
from django.contrib.auth import authenticate, login
from .forms import UbicacionForm
from django.core.files.storage import FileSystemStorage
from xhtml2pdf import pisa
import openpyxl
from openpyxl.chart import PieChart, Reference
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.conf import settings
from .models import Ubicacion, UbicacionCategoria, CategoriaResiduo, Carrito, RegistroResiduo, Rol, Usuario, Empresa, Evento, Recoleccion, Factura, DetalleFactura
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives
def home1(request):
    return render(request, 'home.html')

def home(request):
    return render(request, 'home.html')

def logout_view(request):
    logout(request)  # Cierra la sesión del usuario
    return redirect(reverse('login'))  # Redirige a la página de login

from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from .models import Usuario
def  home_view(request):
    return render(request, 'home.html', {})

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
def registrar_gestor(request):
    if request.method == 'POST':
        form = RegistroGestorForm(request.POST)
        if form.is_valid():
            usuario = form.save(commit=False)
            # Asignar el rol "Gestor de Residuos"
            rol_gestor, _ = Rol.objects.get_or_create(nombre="Gestor de Residuos")
            usuario.rol = rol_gestor
            usuario.save()
            return redirect('login')  # O redirige donde prefieras
    else:
        form = RegistroGestorForm()
    return render(request, 'registro_gestor.html', {'form': form})

def login_usuario(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('username')  # AuthenticationForm usa 'username' por defecto, aunque sea email
            password = form.cleaned_data.get('password')
            user = authenticate(email=email, password=password)  # Autenticamos con email

            if user is not None:
                login(request, user)

                # Redirigir según el rol del usuario
                rol_nombre = user.rol.nombre if user.rol else ''
                rutas_roles = {
                    'Gestor de Residuos': 'vista_gestor_residuos',
                    'experto_ambiental': 'vista_experto_ambiental',
                    'clienteempresa': 'vista_clienteempresa',
                    'clientenatural': 'vista_clientenatural'
                }
                return redirect(rutas_roles.get(rol_nombre, 'home'))  # Si no tiene rol, va a 'home'

            else:
                form.add_error(None, 'Correo o contraseña incorrectos')

    else:
        form = AuthenticationForm()

    return render(request, 'registration/login.html', {'form': form})

@login_required
def vista_gestor_residuos(request):
    return render(request, 'gestor_residuos/gestor_residuos.html')

@login_required
def vista_experto_ambiental(request):
    return render(request, 'experto_ambiental.html')

@login_required
def vista_clienteempresa(request):
    return render(request, 'cliente_empresa/cliente_empresa.html')

@login_required
def vista_clientenatural(request):
    return render(request, 'cliente_natural/cliente_natural.html')

def logout_usuario(request):
    logout(request)
    return redirect('login_usuario')

    

def cliente_empresa_dashboard(request):
    return render(request, 'cliente_empresa/cliente_empresa.html')

def cliente_natural(request):
    return render(request, 'cliente_natural/cliente_natural.html')

def gestor_dashboard(request):

    return render(request, 'gestor_dashboard.html')


#mapa

from django.core.serializers import serialize
from django.shortcuts import render
from .models import Ubicacion

def mapa_ubicaciones(request):
    ubicaciones = Ubicacion.objects.all()
    ubicaciones_json = serialize('json', ubicaciones, fields=['id', 'nombre', 'latitud', 'longitud', 'descripcion'])
    print(ubicaciones_json)  # Depuración para verificar el contenido
    return render(request, 'mapa_ubicaciones.html', {'ubicaciones_json': ubicaciones_json})


#pago
stripe.api_key = settings.STRIPE_SECRET_KEY

def payment_view(request):
    return render(request, 'payment.html', {
        'stripe_public_key': settings.STRIPE_PUBLIC_KEY
    })


@csrf_exempt
def create_checkout_session(request):
    if request.method == 'POST':
        try:
            # Crear una sesión de pago de Stripe
            checkout_session = stripe.checkout.Session.create(
                payment_method_types=['card'],
                line_items=[
                    {
                        'price_data': {
                            'currency': 'usd',
                            'product_data': {
                                'name': 'Nombre del producto',
                            },
                            'unit_amount': 2000,  # Precio en centavos, ejemplo: $20.00
                        },
                        'quantity': 1,
                    },
                ],
                mode='payment',
                success_url='http://localhost:8000/success/',
                cancel_url='http://localhost:8000/cancel/',
            )
            return JsonResponse({
                'id': checkout_session.id
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=403)
def registrar_usuario(request):
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)  # Autentica al usuario después de registrarse
            return redirect('home')  # Redirige a la página principal (ajústalo según tu proyecto)
    else:
        form = RegistroUsuarioForm()
    
    return render(request, 'registro.html', {'form': form})

# Crear un nuevo evento
def crear_evento(request):
    if request.method == 'POST':
        form = EventoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_eventos')
    else:
        form = EventoForm()
    return render(request, 'eventos/crear_evento.html', {'form': form})

# Listar todos los eventos
def listar_eventos(request):
    eventos = Evento.objects.all()
    return render(request, 'eventos/listar_eventos.html', {'eventos': eventos})

# Editar un evento
def editar_evento(request, id):
    evento = get_object_or_404(Evento, id=id)
    if request.method == 'POST':
        form = EventoForm(request.POST, instance=evento)
        if form.is_valid():
            form.save()
            return redirect('listar_eventos')
    else:
        form = EventoForm(instance=evento)
    return render(request, 'eventos/editar_evento.html', {'form': form})

# Eliminar un evento
def eliminar_evento(request, id):
    evento = get_object_or_404(Evento, id=id)
    if request.method == 'POST':
        evento.delete()
        return redirect('listar_eventos')   
    return render(request, 'eventos/eliminar_evento.html', {'evento': evento})

@login_required
def ver_perfil(request):
    usuario = get_object_or_404(Usuario, username=request.user.username)
    return render(request, 'ver_perfil.html', {'usuario': usuario})

@login_required
def editar_perfil(request):
    usuario = get_object_or_404(Usuario, username=request.user.username)
    usuario_residuos = usuario.usuario_residuos

    if request.method == 'POST':
        user_form = UsuarioForm(request.POST, instance=usuario)
        residuos_form = UsuarioResiduosForm(request.POST, instance=usuario_residuos)
        
        if user_form.is_valid() and residuos_form.is_valid():
            user_form.save()
            residuos_form.save()
            return redirect('ver_perfil')  # Redirige a la vista de perfil
    else:
        user_form = UsuarioForm(instance=usuario)
        residuos_form = UsuarioResiduosForm(instance=usuario_residuos)

    return render(request, 'editar_perfil.html', {
        'user_form': user_form,
        'residuos_form': residuos_form
    })


def crear_ubicacion(request):
    if request.method == 'POST':
        ubicacion_form = UbicacionForm(request.POST)
        formset = UbicacionCategoriaFormSet(request.POST)

        if ubicacion_form.is_valid() and formset.is_valid():
            ubicacion = ubicacion_form.save(commit=False)
            ubicacion.usuario = request.user  # Asigna el usuario actual
            ubicacion.save()

            formset.instance = ubicacion
            formset.save()
            return redirect('mapa_ubicaciones1')  # Redirige después de guardar
    else:
        ubicacion_form = UbicacionForm()
        formset = UbicacionCategoriaFormSet()

    return render(request, 'crear_ubicacion.html', {
        'ubicacion_form': ubicacion_form,
        'formset': formset,
    })

#filtro categorias
def ubicacion_list(request):
    ubicaciones = Ubicacion.objects.all()
    return render(request, 'ubicaciones/ubicacion_list.html', {'ubicaciones': ubicaciones})

def ubicacion_list1(request):
    ubicaciones = Ubicacion.objects.all()
    return render(request, 'ubicaciones/ubicacion_list1.html', {'ubicaciones': ubicaciones})
from .forms import BuscarCategoriaForm

@login_required
def mis_ubicaciones(request):
    ubicaciones = Ubicacion.objects.filter(usuario=request.user)
    return render(request, 'mis_ubicaciones.html', {'ubicaciones': ubicaciones})

@login_required
def actualizar_ubicacion(request, pk):
    ubicacion = get_object_or_404(Ubicacion, pk=pk, usuario=request.user)
    if request.method == 'POST':
        form = UbicacionForm(request.POST, instance=ubicacion)
        if form.is_valid():
            form.save()
            return redirect('mis_ubicaciones')
    else:
        form = UbicacionForm(instance=ubicacion)
    return render(request, 'ubicaciones/editar_ubicacion.html', {'form': form})

@login_required
def eliminar_ubicacion(request, pk):
    ubicacion = get_object_or_404(Ubicacion, pk=pk, usuario=request.user)
    if request.method == 'POST':
        ubicacion.delete()
        return redirect('mis_ubicaciones')
    return render(request, 'ubicaciones/confirmar_eliminacion.html', {'ubicacion': ubicacion})


def buscar_por_categoria(request):
    ubicaciones = None
    form = BuscarCategoriaForm(request.GET)
    
    if form.is_valid():
        categoria = form.cleaned_data.get('categoria')
        if categoria:
            ubicaciones = Ubicacion.objects.filter(categoria__icontains=categoria)  # Filtra por categoría

    return render(request, 'ubicaciones/buscar.html', {'form': form, 'ubicaciones': ubicaciones})
@login_required
def cotizar_ubicacion(request, ubicacion_id):
    ubicacion = get_object_or_404(Ubicacion, id=ubicacion_id)
    categorias_disponibles = UbicacionCategoria.objects.filter(ubicacion=ubicacion)
 
    if request.method == 'POST':
        for ubicacion_categoria in categorias_disponibles:
            cantidad_str = request.POST.get(f'cantidad_{ubicacion_categoria.categoria.id}')
            if cantidad_str:
                try:
                    cantidad = float(cantidad_str)
                    if cantidad > 0:
                        Carrito.objects.create(
                            usuario=request.user,
                            ubicacion_categoria=ubicacion_categoria,
                            categoria=ubicacion_categoria.categoria,
                            cantidad_kg=cantidad
                        )
                except ValueError:
                    pass  # Ignorar valores no numéricos

        return redirect('ver_carrito')

    return render(request, 'cotizaciones/cotizar_domicilio.html', {
        'ubicacion': ubicacion,
        'categorias_disponibles': categorias_disponibles,
    })

#expe
@login_required
def agregar_informe_manejo_residuos(request):
    """
    Vista para agregar un informe de manejo de residuos.
    """
    if request.method == 'POST':
        form = InformeManejoResiduosForm(request.POST, request.FILES)
        if form.is_valid():
            descripcion = form.cleaned_data['descripcion']
            archivo = form.cleaned_data['archivo_informe']
            
            file_url = None
            
            # Guardar archivo en el sistema de archivos
            if archivo:
                fs = FileSystemStorage()
                filename = fs.save(archivo.name, archivo)
                file_url = fs.url(filename)
            
            # Confirmación de éxito
            return render(request, 'informes/informe_exitoso.html', {
                'descripcion': descripcion,
                'file_url': file_url,
            })
    else:
        form = InformeManejoResiduosForm()

    return render(request, 'informes/agregar_informe.html', {'form': form, 'titulo': 'Informe de Manejo de Residuos'})


@login_required
def agregar_informe_normativas(request):
    """
    Vista para agregar un informe de normativas.
    """
    if request.method == 'POST':
        form = InformeNormativasForm(request.POST, request.FILES)
        if form.is_valid():
            descripcion = form.cleaned_data['descripcion']
            archivo = form.cleaned_data['archivo_informe']
            
            # Guardar archivo en el sistema de archivos
            if archivo:
                fs = FileSystemStorage()
                filename = fs.save(archivo.name, archivo)
                file_url = fs.url(filename)
            
            # Confirmación de éxito
            return render(request, 'informes/informe_exitoso.html', {
                'descripcion': descripcion,
                'file_url': file_url,
            })
    else:
        form = InformeNormativasForm()

    return render(request, 'informes/agregar_informe.html', {'form': form, 'titulo': 'Informe de Normativas'})

from django.shortcuts import render
from .models import RegistroResiduo

def resumen_co2(request):
    registros = RegistroResiduo.objects.select_related('categoria')
    
    co2_por_categoria = registros.values('categoria__nombre').annotate(
        total_co2=Sum(
            ExpressionWrapper(
                F('cantidad_kg') * F('categoria__factor_co2'),
                output_field=FloatField()
            )
        )
    )

    total_co2 = sum(cat['total_co2'] for cat in co2_por_categoria)

    return render(request, 'resumen_co2.html', {
        'co2_por_categoria': co2_por_categoria,
        'total_co2': total_co2
    })

def lista_categorias(request):
    categorias = CategoriaResiduo.objects.all()
    return render(request, 'categorias/lista.html', {'categorias': categorias})

# CREAR
def crear_categoria(request):
    if request.method == 'POST':
        form = CategoriaResiduoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_categorias')
    else:
        form = CategoriaResiduoForm()
    return render(request, 'categorias/formulario.html', {'form': form, 'accion': 'Crear'})

# ACTUALIZAR
def editar_categoria(request, pk):
    categoria = get_object_or_404(CategoriaResiduo, pk=pk)
    if request.method == 'POST':
        form = CategoriaResiduoForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            return redirect('lista_categorias')
    else:
        form = CategoriaResiduoForm(instance=categoria)
    return render(request, 'categorias/formulario.html', {'form': form, 'accion': 'Editar'})

# ELIMINAR
def eliminar_categoria(request, pk):
    categoria = get_object_or_404(CategoriaResiduo, pk=pk)
    if request.method == 'POST':
        categoria.delete()
        return redirect('lista_categorias')
    return render(request, 'categorias/confirmar_eliminar.html', {'categoria': categoria})

from django.shortcuts import redirect, get_object_or_404
from .models import CategoriaResiduo, Carrito

def añadir_al_carrito(request):
    if request.method == "POST":
        # Verificar si el usuario está autenticado
        if not request.user.is_authenticated:
            return redirect('login')  # Redirigir al login si el usuario no está autenticado
        
        # Iterar sobre cada campo de cantidad de los formularios enviados
        for key, value in request.POST.items():
            if key.startswith('cantidad_kg_'):
                # Obtener la categoría de residuo
                categoria_id = key.split('_')[-1]
                categoria = get_object_or_404(CategoriaResiduo, id=categoria_id)
                
                # Verificar que la cantidad sea un número positivo
                try:
                    cantidad = float(value)
                    if cantidad > 0:
                        # Crear o actualizar la entrada en el carrito
                        carrito_item, created = Carrito.objects.get_or_create(
                            usuario=request.user,
                            categoria=categoria
                        )
                        # Actualizar la cantidad en el carrito
                        carrito_item.cantidad_kg += cantidad
                        carrito_item.save()
                except ValueError:
                    # Si no se puede convertir a float, ignorar ese valor
                    pass
    
    return redirect('ver_carrito')  # Redirigir a la página donde el usuario puede ver el carrito

@login_required
def ver_carrito(request):
    carrito_items = Carrito.objects.filter(usuario=request.user)
    return render(request, 'ver_carrito.html', {'carrito_items': carrito_items})

def mapa_ubicaciones2(request):
    ubicaciones = Ubicacion.objects.all()
    
    ubicaciones_data = []
    for ubicacion in ubicaciones:
        ubicaciones_data.append({
            'id': ubicacion.id, 
            'nombre': ubicacion.nombre,
            'latitud': ubicacion.latitud,
            'longitud': ubicacion.longitud,
            
            
        })
        
    context = {
        'ubicaciones_json': json.dumps(ubicaciones_data)
    }
    return render(request, 'mapa_ubicaciones2.html', context)


def mapa_ubicaciones1(request):
    ubicaciones = Ubicacion.objects.all()
    
    ubicaciones_data = []
    for ubicacion in ubicaciones:
        ubicaciones_data.append({
            'id': ubicacion.id, 
            'nombre': ubicacion.nombre,
            'latitud': ubicacion.latitud,
            'longitud': ubicacion.longitud,
            
            
        })
        
    context = {
        'ubicaciones_json': json.dumps(ubicaciones_data)
    }
    return render(request, 'mapa_ubicaciones1.html', context)
def agregar_empresa(request):
    if request.method == 'POST':






        form = EmpresaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('mapa_ubicaciones')  # Redirige al mapa después de crear la empresa
    else:
        form = EmpresaForm()
    return render(request, 'templates/agregar_empresa.html', {'form': form})

@login_required
def enviar_cotizacion(request):
    carrito_items = Carrito.objects.filter(usuario=request.user)
    if not carrito_items:
        return redirect('carrito_vacio')

    total_compra = sum(item.precio_total for item in carrito_items)

    # Obtener la ubicación y su creador desde el primer ítem
    ubicacion = carrito_items.first().ubicacion_categoria.ubicacion
    creador_ubicacion = ubicacion.usuario

    # Crear la factura con estado pendiente
    factura = Factura.objects.create(
        usuario=request.user,
        creador_ubicacion=creador_ubicacion,
        total=total_compra,
        estado='pendiente'
    )

    for item in carrito_items:
        DetalleFactura.objects.create(
            factura=factura,
            categoria=item.categoria,
            cantidad_kg=item.cantidad_kg,
            precio_unitario=item.precio_unitario,
            precio_total=item.precio_total
        )

    carrito_items.delete()

    messages.success(request, f"Se envió la cotización al dueño de la ubicación: {creador_ubicacion.username}")
    return redirect('direccion_recoleccion', factura_id=factura.id)

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Factura, Recoleccion
from .forms import RecoleccionForm

@login_required
def direccion_recoleccion(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id, usuario=request.user)

    # Evita que se registre más de una recolección para una misma factura
    if hasattr(factura, 'recoleccion'):
        messages.warning(request, "Ya has registrado una dirección de recolección para esta factura.")
        return redirect('mapa_ubicaciones2')  # O alguna otra vista relevante

    if request.method == 'POST':
        form = RecoleccionForm(request.POST)
        if form.is_valid():
            recoleccion = form.save(commit=False)
            recoleccion.factura = factura
            recoleccion.save()

            messages.success(request, "La dirección fue registrada. El recolector ha sido notificado.")
            return redirect('mapa_ubicaciones2')
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = RecoleccionForm()

    return render(request, 'direccion_form.html', {
        'form': form,
        'factura': factura
    })


def generar_pdf(factura, co2_acumulado):
    # Renderizar el HTML para la factura
    html = render_to_string('factura_pdf.html', {
        'factura': factura,
        'co2_acumulado': round(co2_acumulado, 2),
    })

    # Convertir HTML a PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="factura_{}.pdf"'.format(factura.id)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    
    return response
@login_required
def historial_facturas(request):
    facturas = Factura.objects.filter(usuario=request.user).order_by('-fecha')

    return render(request, 'historial_facturas.html', {
        'facturas': facturas
    })

def pago_exitoso(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)

    # Cálculo del CO₂ evitado acumulado
    co2_acumulado = RegistroResiduo.objects.annotate(
        co2=ExpressionWrapper(
            F('cantidad_kg') * F('categoria__factor_co2'),
            output_field=FloatField()
        )
    ).aggregate(total_co2=Sum('co2'))['total_co2'] or 0

    return render(request, 'pago_exitoso.html', {
        'factura': factura,
        'co2_acumulado': round(co2_acumulado, 2),
    })


def carrito_vacio(request):
    return render(request, 'carrito_vacio.html')

import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import RegistroResiduo
from collections import defaultdict

def exportar_excel(request):
    registros = RegistroResiduo.objects.select_related('categoria')

    # Agrupar por categoría
    datos = defaultdict(lambda: {'cantidad_kg': 0, 'co2': 0})
    for r in registros:
        nombre = r.categoria.nombre
        datos[nombre]['cantidad_kg'] += r.cantidad_kg
        datos[nombre]['co2'] += r.co2_evitable

    # Crear libro
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen CO2"

    # Encabezado
    encabezados = ["Categoría", "Cantidad (kg)", "CO₂ Evitado (kg)"]
    ws.append(encabezados)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Datos
    for nombre, valores in datos.items():
        ws.append([nombre, valores['cantidad_kg'], valores['co2']])

    # Crear gráfico de pastel
    chart = PieChart()
    chart.title = "Distribución de CO₂ Evitado por Categoría"
    data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    ws.add_chart(chart, "E5")  # Posición del gráfico

    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="resumen_co2.xlsx"'
    wb.save(response)
    return response

def exportar_pdf(request):
    registros = RegistroResiduo.objects.select_related('categoria')
    total_co2 = sum(r.co2_evitable for r in registros)

    template_path = 'reporte_co2_pdf.html'
    context = {
        'registros': registros,
        'total_co2': total_co2
    }

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="resumen_co2.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

def buscar_por_categoria(request):
    ubicaciones = None
    form = BuscarCategoriaForm(request.GET or None)

    if form.is_valid():
        categoria = form.cleaned_data['categoria']
        ubicaciones = Ubicacion.objects.filter(categorias=categoria).distinct()

    return render(request, 'buscar_por_categoria.html', {
        'form': form,
        'ubicaciones': ubicaciones
    })
@login_required
def ver_prefactura(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id, creador_ubicacion=request.user)
    return render(request, 'cotizaciones/prefactura.html', {'factura': factura})
@login_required
def ver_cotizaciones_recibidas(request):
    facturas = Factura.objects.filter(creador_ubicacion=request.user).prefetch_related('detalles', 'detalles__categoria').order_by('-fecha')
    return render(request, 'cotizaciones/recibidas.html', {'facturas': facturas})

#@login_required
#def cambiar_estado(request, factura_id, nuevo_estado):
#    factura = get_object_or_404(Factura, id=factura_id)
    # Estados válidos definidos en el modelo
 #   estados_validos = [choice[0] for choice in Factura.ESTADOS]

  #  if nuevo_estado not in estados_validos:
   #     messages.error(request, "Estado no válido.")
    #else:
     #   factura.estado = nuevo_estado
      #  factura.save()
       # messages.success(request, f"Estado actualizado a «{factura.get_estado_display()}».")

    #return redirect('ver_cotizaciones_recibidas')

@login_required
def cambiar_estado(request, factura_id, nuevo_estado):
    factura = get_object_or_404(Factura, id=factura_id)
    estados_validos = [choice[0] for choice in Factura.ESTADOS]

    if nuevo_estado not in estados_validos:
        messages.error(request, "Estado no válido.")
    else:
        factura.estado = nuevo_estado
        factura.save()
        messages.success(request, f"Estado actualizado a «{factura.get_estado_display()}».")

        # Enviar correo si el estado es "recolectado"
        if nuevo_estado == 'recolectado':
            cliente = factura.usuario

            # Renderizar plantilla HTML del correo
            subject = '✅ Recolección Completada - Ecoresiduos'
            from_email = settings.DEFAULT_FROM_EMAIL
            to = [cliente.email]

            context = {
                'nombre': cliente.get_full_name() or cliente.username,
                'factura': factura,
            }

            text_content = f'Hola, tu recolección ha sido completada.'
            html_content = render_to_string('emails/recoleccion_completada.html', context)

            msg = EmailMultiAlternatives(subject, text_content, from_email, to)
            msg.attach_alternative(html_content, "text/html")
            msg.send()

    return redirect('ver_cotizaciones_recibidas')

def ver_ruta(request, factura_id):
    recoleccion = get_object_or_404(Recoleccion, factura_id=factura_id)
    return render(request, 'cotizaciones/ver_ruta.html', {
        'recoleccion': recoleccion,
        'coordenadas': recoleccion.coordenadas  # (lat, lng) o (None, None)
    })
@login_required
#def marcar_recolectado(request, factura_id):
 #   factura = get_object_or_404(Factura, id=factura_id, creador_ubicacion=request.user)
#
 #   if factura.estado != 'recolectado':
  #      factura.estado = 'recolectado'
   #     factura.save()
#
        # Registrar residuos recolectados por categoría
 #       for detalle in factura.detalles.all():
  #         RegistroResiduo.objects.create(
   #             categoria=detalle.categoria,
    #            cantidad_kg=detalle.cantidad_kg
  #          )
#
 #       messages.success(request, "La recolección fue marcada como completada y los residuos fueron registrados.")
   # else:
   #     messages.info(request, "Esta factura ya estaba marcada como recolectada.")

    #return redirect('cotizaciones_recibidas')

@login_required
def marcar_recolectado(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)
    factura.estado = 'en_recoleccion'
    factura.save()
    
    cliente = factura.usuario
    send_mail(
        'Recolección Completada',
        f'Hola {cliente.username},\n\nLa recolección de residuos ha sido programada.',
        settings.DEFAULT_FROM_EMAIL,
        [cliente.email],
        fail_silently=False,
    )
    
    return redirect('ver_cotizaciones_recibidas')